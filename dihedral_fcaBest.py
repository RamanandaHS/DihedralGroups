# Research-grade implementation using PURE NUMERIC ENCODING
# Objects and attributes are represented as lists:
#
# Object encoding:
#   Reflection <sr^j>        -> [0, j]
#   Cyclic subgroup <r^d>    -> [d, 0]
#
# Attribute encoding:
#   Dihedral <r^d, sr^j>     -> [d, j]
#   Full rotation <r>        -> [1, 0]
#
# No strings are used in incidence logic.

from openpyxl import Workbook
from sympy import factorint
import csv

# ------------------------------------------------------------
# Utilities
# ------------------------------------------------------------

def divisors(n):
    return [d for d in range(1, n+1) if n % d == 0]

def is_prime_power(x):
    f = factorint(x)
    return len(f) == 1


# ------------------------------------------------------------
# Generate Objects
# ------------------------------------------------------------

def generate_objects(n):
    objects = []
    
    # Reflection objects [0, j]
    for j in range(n):
        objects.append([0, j])
    
    # Cyclic objects [d, 0] where n/d is prime power
    for d in divisors(n):
        if n // d > 1 and is_prime_power(n // d):
            objects.append([d, 0])
    
    return objects


# ------------------------------------------------------------
# Generate Attributes
# ------------------------------------------------------------

def generate_attributes(n):
    attributes = []
    factorization = factorint(n)
    
    # Dihedral attributes [d, j] where d = p^k
    for p, alpha in factorization.items():
        for k in range(1, alpha + 1):
            d = p**k
            for j in range(d):
                attributes.append([d, j])
    
    # Full rotation attribute [1, 0]
    attributes.append([1, 0])
    
    return attributes


# ------------------------------------------------------------
# Incidence Function (Pure Arithmetic)
# ------------------------------------------------------------

def incidence(obj, attr):
    
    # Reflection object [0, i]
    if obj[0] == 0:
        i = obj[1]
        
        # If attribute is dihedral [d, j]
        if attr[0] != 0 and attr!=[1,0]:
            d = attr[0]
            j = attr[1]
            return 1 if i % d == j % d else 0
        
        return 0
    
    # Cyclic object [d_obj, 0]
    else:
        d_obj = obj[0]
        
        # Dihedral attribute [d_attr, j]
        d_attr = attr[0]
        
        # Full rotation case [1,0]
        if attr == [1, 0]:
            return 1
        
        # Divisibility condition
        return 1 if d_obj % d_attr == 0 else 0


# ------------------------------------------------------------
# Export to Excel
# ------------------------------------------------------------

def export_context_table(n):
    objects = generate_objects(n)
    attributes = generate_attributes(n)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"D{n}_Context"
    
    ws.cell(row=1, column=1, value="Objects \\ Attributes")
    
    # Write attribute headers
    for col, attr in enumerate(attributes, start=2):
        ws.cell(row=1, column=col, value=str(attr))
    
    # Write objects and matrix
    for row_idx, obj in enumerate(objects, start=2):
        ws.cell(row=row_idx, column=1, value=str(obj))
        for col_idx, attr in enumerate(attributes, start=2):
            ws.cell(row=row_idx, column=col_idx, value=incidence(obj, attr))
    
    filename = f"E:/CodeApps/context/D{n}_table.xlsx"
    wb.save(filename)
    return filename


# ------------------------------------------------------------
# Run
# ------------------------------------------------------------

n = int(input("Enter n for D_{n}: ").strip())
file_path = export_context_table(n)

file_path